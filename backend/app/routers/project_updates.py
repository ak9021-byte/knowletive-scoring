import io
import re
import base64
from datetime import date as date_type, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from database import get_db
from app.models.project_update import ProjectUpdate
from app.models.student import Student

router = APIRouter(prefix="/project-updates", tags=["Project Updates"])


class ProjectUpdateCreate(BaseModel):
    student_id: int
    name: str
    project_name: str
    technology: Optional[str] = None
    date: Optional[str] = None
    time: Optional[str] = None
    image: Optional[str] = None
    github_link: Optional[str] = None
    deployment_link: Optional[str] = None


class ApprovePayload(BaseModel):
    remark: Optional[str] = None
    reviewer_name: Optional[str] = None


@router.post("/")
def create_update(payload: ProjectUpdateCreate, db: Session = Depends(get_db)):
    now = datetime.now()
    update = ProjectUpdate(
        student_id=payload.student_id,
        name=payload.name,
        project_name=payload.project_name,
        technology=payload.technology,
        date=payload.date or date_type.today().isoformat(),
        time=payload.time or now.strftime("%H:%M"),
        image=payload.image,
        github_link=payload.github_link,
        deployment_link=payload.deployment_link,
    )
    db.add(update)
    db.commit()
    db.refresh(update)
    return update


@router.patch("/{update_id}/approve")
def approve_update(update_id: int, payload: ApprovePayload, db: Session = Depends(get_db)):
    update = db.query(ProjectUpdate).filter(ProjectUpdate.id == update_id).first()
    if not update:
        raise HTTPException(status_code=404, detail="Update not found")
    update.approved = True
    update.faculty_remark = payload.remark or "Today's work done ✅"
    update.reviewer_name = payload.reviewer_name
    db.commit()
    db.refresh(update)
    return update


@router.get("/student/{student_id}")
def get_student_updates(student_id: int, db: Session = Depends(get_db)):
    updates = (
        db.query(ProjectUpdate)
        .filter(ProjectUpdate.student_id == student_id)
        .order_by(ProjectUpdate.created_at.desc())
        .all()
    )
    return updates


@router.get("/all")
def get_all_updates(db: Session = Depends(get_db)):
    results = (
        db.query(ProjectUpdate)
        .order_by(ProjectUpdate.created_at.desc())
        .all()
    )
    return results


def _safe_sheet_name(name: str, used: set) -> str:
    clean = re.sub(r'[:\\/?*\[\]]', "", name).strip() or "Student"
    clean = clean[:31]
    base = clean
    i = 2
    while clean in used:
        suffix = f" ({i})"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


def _write_sheet(ws, rows):
    headers = ["Date", "Time", "Project Name", "Technology", "GitHub Link", "Deployment Link", "Faculty Remark", "Reviewed By", "Image"]
    ws.append(headers)
    widths = [14, 10, 22, 20, 32, 32, 26, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    row_idx = 2
    for u in rows:
        ws.cell(row=row_idx, column=1, value=u.date)
        ws.cell(row=row_idx, column=2, value=u.time)
        ws.cell(row=row_idx, column=3, value=u.project_name)
        ws.cell(row=row_idx, column=4, value=u.technology or "")
        ws.cell(row=row_idx, column=5, value=u.github_link or "")
        ws.cell(row=row_idx, column=6, value=u.deployment_link or "")
        ws.cell(row=row_idx, column=7, value=(u.faculty_remark if u.approved else "Pending review"))
        ws.cell(row=row_idx, column=8, value=(u.reviewer_name or "") if u.approved else "")

        if u.image:
            try:
                b64data = u.image.split(",", 1)[1] if "," in u.image else u.image
                img_bytes = base64.b64decode(b64data)
                pil_img = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
                pil_img.thumbnail((140, 140))
                buf = io.BytesIO()
                pil_img.save(buf, format="PNG")
                buf.seek(0)
                xl_img = XLImage(buf)
                ws.add_image(xl_img, f"I{row_idx}")
                ws.row_dimensions[row_idx].height = 105
            except Exception:
                ws.cell(row=row_idx, column=9, value="Image error")
        row_idx += 1


@router.get("/export")
def export_excel(db: Session = Depends(get_db)):
    results = (
        db.query(ProjectUpdate)
        .order_by(ProjectUpdate.name.asc(), ProjectUpdate.date.desc(), ProjectUpdate.created_at.desc())
        .all()
    )

    grouped: dict[str, list] = {}
    for u in results:
        grouped.setdefault(u.name, []).append(u)

    wb = Workbook()
    wb.remove(wb.active)

    used_names: set = set()
    if not grouped:
        ws = wb.create_sheet("No Data")
        ws.append(["No project updates submitted yet"])
    else:
        for student_name in sorted(grouped.keys()):
            sheet_name = _safe_sheet_name(student_name, used_names)
            ws = wb.create_sheet(sheet_name)
            _write_sheet(ws, grouped[student_name])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=project_updates.xlsx"},
    )